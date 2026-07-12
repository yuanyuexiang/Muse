"""菜单内容提取：把一批筛过的消息（文本 + 图片）→ 结构化 MenuSpec。

客户是餐厅老板，要为其设计一张菜单。提取的是**整份菜单的内容**：店铺信息、
分类菜品(编号/名/描述/价/标记)、套餐 —— 这份结构既是审校对象，也是 HTML/CSS
模板渲染成印刷 PDF 的数据源。

Phase 1 就是**一次带结构化输出的多模态 LLM 调用**，不需要 LangGraph。
- 图片(老菜单照片/Excel 截图)走 image_url 内联给 VLM(qwen-vl-max)直读。
- 未配置 LLM_API_KEY 时返回占位草稿，保证管线在无 LLM/无凭证时也能端到端跑通。
"""

import base64
import io
import json
import logging

from app.config import settings
from app.models import InboxMessage
from app.schemas import MenuSpec
from app.storage import storage

log = logging.getLogger("muse.llm")

SYSTEM_PROMPT = (
    "你是餐厅菜单结构化助手。餐厅老板通过客服发来菜品清单/老菜单照片/Excel 等，"
    "你要把**整份菜单的内容**提取成 JSON：店铺信息(店名/标语/电话/地址/营业时间/"
    "促销/过敏原)、所有分类及其菜品(编号 number / 名称 name / 描述 / 价格 price 保留"
    "原样如 £6.00 / 标记 flags: hot|vegetarian|nut)、套餐 set_meals。尽量不漏菜。"
    "拿不准的放进 missing_fields。严格按给定 JSON schema，只输出 JSON。"
)

_MIME = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
         "gif": "image/gif", "webp": "image/webp", "bmp": "image/bmp"}


def _schema_instruction() -> str:
    return "请严格按此 JSON schema 输出，仅输出 JSON：\n" + json.dumps(
        MenuSpec.model_json_schema(), ensure_ascii=False
    )


def _guess_mime(object_key: str) -> str:
    ext = object_key.rsplit(".", 1)[-1].lower() if "." in object_key else ""
    return _MIME.get(ext, "image/jpeg")


def _image_data_url(object_key: str) -> str | None:
    try:
        raw = storage.load(object_key)
    except Exception as exc:  # noqa: BLE001
        log.warning("load image failed key=%s: %s", object_key, exc)
        return None
    mime = _guess_mime(object_key)
    try:
        from PIL import Image

        img = Image.open(io.BytesIO(raw)).convert("RGB")
        img.thumbnail((1900, 1900))
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=88)
        raw, mime = buf.getvalue(), "image/jpeg"
    except Exception as exc:  # noqa: BLE001
        log.warning("image transcode fallback key=%s: %s", object_key, exc)
    return f"data:{mime};base64,{base64.b64encode(raw).decode()}"


def _file_text(object_key: str, max_chars: int = 8000) -> str | None:
    """抽取 Excel/PDF 正文喂给提取。扫描版 PDF(无文本层)返回 None。"""
    ext = object_key.rsplit(".", 1)[-1].lower() if "." in object_key else ""
    try:
        raw = storage.load(object_key)
    except Exception:  # noqa: BLE001
        return None
    try:
        if ext in ("xlsx", "xlsm"):
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            rows = []
            for ws in wb.worksheets:
                rows.append(f"# 工作表：{ws.title}")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        rows.append(" | ".join(cells))
            return "\n".join(rows)[:max_chars] or None
        if ext == "pdf":
            from pypdf import PdfReader

            text = "\n".join((p.extract_text() or "") for p in PdfReader(io.BytesIO(raw)).pages)
            return text[:max_chars] if text.strip() else None
    except Exception as exc:  # noqa: BLE001
        log.warning("file parse failed key=%s: %s", object_key, exc)
    return None


def _file_or_placeholder(m: InboxMessage) -> str:
    if m.type == "file" and m.object_key:
        txt = _file_text(m.object_key)
        if txt:
            return f"[{m.seq}] 文件《{m.object_key.rsplit('/', 1)[-1]}》内容：\n{txt}"
    return f"[{m.seq}] <{m.type}:{m.object_key or ''}>"


def _has_images(messages: list[InboxMessage]) -> bool:
    return any(m.type == "image" and m.object_key for m in messages)


def _text_prompt(messages: list[InboxMessage]) -> str:
    lines = []
    for m in sorted(messages, key=lambda x: x.seq):
        if m.type == "text" or (m.type == "voice" and m.content):
            lines.append(f"[{m.seq}] {m.content or ''}")
        else:
            lines.append(_file_or_placeholder(m))
    return "客户发来的内容：\n" + "\n".join(lines) + "\n\n" + _schema_instruction()


def _multimodal_content(messages: list[InboxMessage]) -> list[dict]:
    parts: list[dict] = [{"type": "text", "text": "客户发来的内容（含图片）："}]
    for m in sorted(messages, key=lambda x: x.seq):
        if m.type == "image" and m.object_key:
            url = _image_data_url(m.object_key)
            if url:
                parts.append({"type": "text", "text": f"[{m.seq}] 图片："})
                parts.append({"type": "image_url", "image_url": {"url": url}})
            else:
                parts.append({"type": "text", "text": f"[{m.seq}] <图片下载缺失>"})
        elif m.type == "text" or (m.type == "voice" and m.content):
            parts.append({"type": "text", "text": f"[{m.seq}] {m.content or ''}"})
        else:
            parts.append({"type": "text", "text": _file_or_placeholder(m)})
    parts.append({"type": "text", "text": _schema_instruction()})
    return parts


async def extract_menu_spec(messages: list[InboxMessage]) -> MenuSpec:
    if not settings.llm_api_key:
        return MenuSpec(
            notes="⚠️ LLM 未配置（占位草稿）。原始消息：\n" + _text_prompt(messages),
            missing_fields=["shop", "categories"],
        )

    from openai import AsyncOpenAI

    client = AsyncOpenAI(base_url=settings.llm_base_url, api_key=settings.llm_api_key)
    user_content = (
        _multimodal_content(messages) if _has_images(messages) else _text_prompt(messages)
    )
    try:
        resp = await client.chat.completions.create(
            model=settings.llm_model,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_content},
            ],
            response_format={"type": "json_object"},
            temperature=0,
        )
        payload = json.loads(resp.choices[0].message.content or "{}")
        return MenuSpec.model_validate(payload)
    except Exception as exc:  # noqa: BLE001 — 提取失败不阻断，交人工审校
        log.warning("menu extraction failed: %s", exc)
        return MenuSpec(notes=f"⚠️ 菜单提取失败：{exc}", missing_fields=["shop", "categories"])
